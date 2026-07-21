#!/usr/bin/env python3
"""Generate supabase/functions/pricing/wageTable.generated.ts from BLS OEWS data.

Downloads the OEWS state and national XLSX files for a given vintage year and
extracts hourly wage percentiles (p25 / median / p75, cross-industry) for the
construction-trade occupations the pricing engine uses. Re-run once a year when
BLS publishes the new vintage (May data ships the following March/April).

Usage:
    python3 scripts/ingest-oews.py [--year 25] [--keep-downloads]

BLS blocks anonymous scripted downloads; per their usage policy the request
identifies itself with a contact email in the User-Agent (BLS_CONTACT env var
overrides the default).

Requires: openpyxl (pip3 install --user openpyxl)
"""

import argparse
import io
import os
import sys
import urllib.request
import zipfile
from datetime import date
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = REPO_ROOT / "supabase" / "functions" / "pricing" / "wageTable.generated.ts"

BLS_BASE = "https://www.bls.gov/oes/special-requests"
CONTACT = os.environ.get("BLS_CONTACT", "hello@brightglow.co")
USER_AGENT = f"Brightglow pricing research ({CONTACT})"

# SOC occupations the cost catalog prices labor against. Keys must match the
# `soc` field in costCatalog.ts entries.
SOC_CODES = {
    "47-2031": "Carpenters",
    "47-2041": "Carpet Installers",
    "47-2042": "Floor Layers, Except Carpet, Wood, and Hard Tiles",
    "47-2043": "Floor Sanders and Finishers",
    "47-2044": "Tile and Stone Setters",
    "47-2111": "Electricians",
    "47-2141": "Painters, Construction and Maintenance",
    "47-2152": "Plumbers, Pipefitters, and Steamfitters",
    "47-2181": "Roofers",
    "49-9021": "Heating, Air Conditioning, and Refrigeration Mechanics and Installers",
    "37-3011": "Landscaping and Groundskeeping Workers",
    "49-9071": "Maintenance and Repair Workers, General",
    # Auto & moto vertical. Shops bill a posted "door rate" rather than a
    # wage, but the door rate tracks the local tech wage the same way a
    # contractor's billed rate does — see AUTO_LABOR_BURDEN in costCatalog.ts.
    "49-3023": "Automotive Service Technicians and Mechanics",
    "49-3021": "Automotive Body and Related Repairers",
    "49-3022": "Automotive Glass Installers and Repairers",
    "49-3052": "Motorcycle Mechanics",
    "53-7061": "Cleaners of Vehicles and Equipment",
}

WAGE_COLS = ("H_PCT25", "H_MEDIAN", "H_PCT75")


def fetch_zip(filename: str) -> bytes:
    url = f"{BLS_BASE}/{filename}"
    print(f"downloading {url} ...", file=sys.stderr)
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req) as res:
        body = res.read()
    print(f"  {len(body) / 1e6:.1f} MB", file=sys.stderr)
    return body


def parse_wage(value):
    """OEWS wage cells: number, None, '*' (unavailable), '#' (>= $115/hr)."""
    if isinstance(value, (int, float)):
        return float(value)
    if value == "#":
        return 115.0
    return None


def extract(xlsx_bytes: bytes, want_states: bool):
    """Return {state_or_'US': {soc: (p25, median, p75)}} for cross-industry rows."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    col = {name: header.index(name) for name in
           ("AREA_TYPE", "PRIM_STATE", "NAICS", "OCC_CODE", *WAGE_COLS)}

    out = {}
    for r in rows:
        if r[col["OCC_CODE"]] not in SOC_CODES:
            continue
        if r[col["NAICS"]] != "000000":  # cross-industry rows only
            continue
        area_type = str(r[col["AREA_TYPE"]])
        if want_states and area_type != "2":
            continue
        if not want_states and area_type != "1":
            continue
        key = r[col["PRIM_STATE"]] if want_states else "US"
        wages = tuple(parse_wage(r[col[c]]) for c in WAGE_COLS)
        if any(w is None for w in wages):
            continue  # drop rows with suppressed percentiles; engine falls back to US
        out.setdefault(key, {})[r[col["OCC_CODE"]]] = wages
    return out


def ts_wage_map(data: dict) -> str:
    lines = []
    for soc in sorted(data):
        p25, median, p75 = data[soc]
        lines.append(f'    "{soc}": {{ p25: {p25}, median: {median}, p75: {p75} }},')
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", default="25", help="two-digit OEWS vintage year (e.g. 25 for May 2025)")
    args = ap.parse_args()

    nat = extract(read_xlsx_from_zip(fetch_zip(f"oesm{args.year}nat.zip")), want_states=False)
    states = extract(read_xlsx_from_zip(fetch_zip(f"oesm{args.year}st.zip")), want_states=True)

    if "US" not in nat:
        sys.exit("national file yielded no cross-industry rows — column layout changed?")
    missing = [s for s in SOC_CODES if s not in nat["US"]]
    if missing:
        sys.exit(f"national data missing SOC codes {missing} — check SOC_CODES against the vintage")

    vintage = f"May 20{args.year}"
    state_blocks = []
    for state in sorted(states):
        state_blocks.append(f'  "{state}": {{\n{ts_wage_map(states[state])}\n  }},')

    OUTPUT.write_text(f"""// AUTO-GENERATED by scripts/ingest-oews.py on {date.today().isoformat()} — do not edit.
// Source: BLS OEWS {vintage} estimates (oesm{args.year}nat.zip / oesm{args.year}st.zip),
// hourly wage percentiles, cross-industry, by state and SOC occupation.
// Public U.S. government data (bls.gov/oes). Re-run the script yearly.

export interface WageRow {{
  p25: number;
  median: number;
  p75: number;
}}

export const WAGE_VINTAGE = "{vintage}";

/** National fallback used when a state row is suppressed in the OEWS release. */
export const NATIONAL_WAGES: Record<string, WageRow> = {{
{ts_wage_map(nat["US"])}
}};

export const STATE_WAGES: Record<string, Record<string, WageRow>> = {{
{chr(10).join(state_blocks)}
}};
""")
    n_rows = sum(len(v) for v in states.values())
    print(f"wrote {OUTPUT.relative_to(REPO_ROOT)}: {len(states)} areas, {n_rows} state rows, vintage {vintage}")


def read_xlsx_from_zip(zip_bytes: bytes) -> bytes:
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    names = [n for n in zf.namelist() if n.endswith(".xlsx") and "dl" in n.lower()]
    if not names:
        names = [n for n in zf.namelist() if n.endswith(".xlsx")]
    return zf.read(names[0])


if __name__ == "__main__":
    main()
