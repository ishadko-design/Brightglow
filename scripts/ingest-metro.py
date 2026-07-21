#!/usr/bin/env python3
"""Generate supabase/functions/pricing/metroWages.generated.ts from BLS OEWS
metropolitan wage data + a Census-derived ZIP→metro crosswalk.

The state-level engine (ingest-oews.py / wageTable.generated.ts) prices San
Francisco and rural California identically. This adds metro (CBSA) resolution:
a zip that falls in an OEWS-covered metro prices on that metro's wages; every
other zip falls back to state wages. Fully automated from three ungated public
files, so it re-runs yearly alongside ingest-oews.py:

  - BLS OEWS metro wages     oesm{YY}ma.zip           (bls.gov)
  - Census ZCTA→county rel.  tab20_zcta520_county20_natl.txt
  - Census CBSA delineation  list1_2023.xlsx          (county→CBSA)

Usage:  python3 scripts/ingest-metro.py [--year 25]
Requires: openpyxl.  BLS blocks anonymous bots — request identifies itself via
BLS_CONTACT (default hello@brightglow.co) in the User-Agent, same as ingest-oews.
"""

import argparse
import csv
import io
import os
import sys
import urllib.request
import zipfile
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = REPO_ROOT / "supabase" / "functions" / "pricing" / "metroWages.generated.ts"

CONTACT = os.environ.get("BLS_CONTACT", "hello@brightglow.co")
USER_AGENT = f"Brightglow pricing research ({CONTACT})"

# Keep in sync with ingest-oews.py SOC_CODES.
SOC_CODES = {
    "47-2031", "47-2041", "47-2042", "47-2043", "47-2044", "47-2111",
    "47-2141", "47-2152", "47-2181", "49-9021", "37-3011", "49-9071",
    "49-3023", "49-3021", "49-3022", "49-3052", "53-7061",
}
WAGE_COLS = ("H_PCT25", "H_MEDIAN", "H_PCT75")

OEWS_MA = "https://www.bls.gov/oes/special-requests/oesm{yy}ma.zip"
ZCTA_COUNTY = "https://www2.census.gov/geo/docs/maps-data/data/rel2020/zcta520/tab20_zcta520_county20_natl.txt"
CBSA_DELIN = "https://www2.census.gov/programs-surveys/metro-micro/geographies/reference-files/2023/delineation-files/list1_2023.xlsx"


def fetch(url: str) -> bytes:
    print(f"downloading {url} ...", file=sys.stderr)
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req) as res:
        body = res.read()
    print(f"  {len(body) / 1e6:.1f} MB", file=sys.stderr)
    return body


def parse_wage(v):
    if isinstance(v, (int, float)):
        return float(v)
    return 115.0 if v == "#" else None


def metro_wages(xlsx_bytes: bytes):
    """{cbsa: {soc: (p25, median, p75)}} from the OEWS MSA file, cross-industry."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    h = list(next(rows))
    ci = {c: h.index(c) for c in ("AREA", "NAICS", "OCC_CODE", *WAGE_COLS)}
    out = defaultdict(dict)
    for r in rows:
        if r[ci["OCC_CODE"]] not in SOC_CODES or r[ci["NAICS"]] != "000000":
            continue
        wages = tuple(parse_wage(r[ci[c]]) for c in WAGE_COLS)
        if any(w is None for w in wages):
            continue
        out[str(r[ci["AREA"]])][r[ci["OCC_CODE"]]] = wages
    return out


def county_to_cbsa(xlsx_bytes: bytes):
    """{county_fips: cbsa} for Metropolitan CBSAs only (micros price on state)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = rows[2]
    cbsaI = h.index("CBSA Code")
    typeI = h.index("Metropolitan/Micropolitan Statistical Area")
    fs, fc = h.index("FIPS State Code"), h.index("FIPS County Code")
    out = {}
    for r in rows[3:]:
        if not r[cbsaI] or r[fs] is None or "Metropolitan" not in str(r[typeI] or ""):
            continue
        out[f"{int(r[fs]):02d}{int(r[fc]):03d}"] = str(r[cbsaI])
    return out


def zcta_to_county(txt_bytes: bytes):
    """{zcta: dominant county_fips} — the county holding the most ZCTA land."""
    best = {}
    f = io.StringIO(txt_bytes.decode("utf-8-sig"))
    r = csv.reader(f, delimiter="|")
    h = next(r)
    zi, ci, ai = (h.index("GEOID_ZCTA5_20"), h.index("GEOID_COUNTY_20"),
                  h.index("AREALAND_PART"))
    for row in r:
        z = row[zi].strip()
        if not z:
            continue
        land = float(row[ai] or 0)
        if z not in best or land > best[z][1]:
            best[z] = (row[ci], land)
    return {z: v[0] for z, v in best.items()}


def zip3_to_cbsa(zcta_county, county_cbsa, oews_cbsas):
    """zip3 prefix → the OEWS-covered metro CBSA most of its ZCTAs fall in.
    A zip3 whose ZCTAs are mostly rural/micro (no OEWS metro) maps to nothing
    and the engine uses state wages."""
    votes = defaultdict(Counter)
    for zcta, county in zcta_county.items():
        cbsa = county_cbsa.get(county)
        if cbsa and cbsa in oews_cbsas:
            votes[zcta[:3]][cbsa] += 1
    out = {}
    for z3, counter in votes.items():
        cbsa, _ = counter.most_common(1)[0]
        out[z3] = cbsa
    return out


def read_ma_xlsx(zip_bytes: bytes) -> bytes:
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    name = next(n for n in zf.namelist() if n.endswith("MSA_M2025_dl.xlsx")
                or (n.endswith(".xlsx") and "MSA" in n))
    return zf.read(name)


def ts_metro_block(wages) -> str:
    lines = []
    for cbsa in sorted(wages):
        socs = wages[cbsa]
        inner = ", ".join(
            f'"{soc}": {{ p25: {socs[soc][0]}, median: {socs[soc][1]}, p75: {socs[soc][2]} }}'
            for soc in sorted(socs)
        )
        lines.append(f'  "{cbsa}": {{ {inner} }},')
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", default="25")
    args = ap.parse_args()

    wages = metro_wages(read_ma_xlsx(fetch(OEWS_MA.format(yy=args.year))))
    county_cbsa = county_to_cbsa(fetch(CBSA_DELIN))
    zcta_county = zcta_to_county(fetch(ZCTA_COUNTY))
    z3 = zip3_to_cbsa(zcta_county, county_cbsa, set(wages.keys()))

    if not wages or not z3:
        sys.exit("empty metro wages or crosswalk — a source layout changed?")

    zip3_lines = "\n".join(f'  "{k}": "{z3[k]}",' for k in sorted(z3))
    vintage = f"May 20{args.year}"
    OUTPUT.write_text(f"""// AUTO-GENERATED by scripts/ingest-metro.py on {date.today().isoformat()} — do not edit.
// BLS OEWS {vintage} metropolitan wages (oesm{args.year}ma.zip) keyed by CBSA,
// plus a zip3→CBSA crosswalk composed from the Census ZCTA→county relationship
// file and the 2023 CBSA delineation. A zip3 maps to the OEWS-covered metro
// most of its ZCTAs fall in; zip3s that are mostly rural/micropolitan are
// absent here and price on state wages (wageTable.generated.ts). All public
// U.S. government data. Re-run yearly with ingest-oews.py.

import type {{ WageRow }} from "./wageTable.generated.ts";

/** OEWS metro wages by CBSA code → SOC occupation. {len(wages)} metros. */
export const METRO_WAGES: Record<string, Record<string, WageRow>> = {{
{ts_metro_block(wages)}
}};

/** 3-digit ZIP prefix → CBSA code, metros with OEWS data only. {len(z3)} prefixes. */
export const ZIP3_CBSA: Record<string, string> = {{
{zip3_lines}
}};
""")
    print(f"wrote {OUTPUT.relative_to(REPO_ROOT)}: {len(wages)} metros, {len(z3)} zip3 prefixes")


if __name__ == "__main__":
    main()
